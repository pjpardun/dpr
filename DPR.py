from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal
from GUI import Ui_MainWindow # generated GUI py file
import sys
import os
import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from datetime import datetime
import ctypes
import matplotlib.pyplot as plt
import pwlf
from GPyOpt.methods import BayesianOptimization
import openpyxl
import math
from scipy import stats


# python included dependencies: sys, os, datetime, dateutil, math
# installed package dependencies: pyqt, pandas (and xlsxwriter), numpy, scipi, matplotlib, pwlf, gpy, openpyxl (and image)

# class to populate a PyQT table view with a pandas dataframe
class PandasModel(QtCore.QAbstractTableModel):

    def __init__(self, data, parent=None):
        QtCore.QAbstractTableModel.__init__(self, parent)
        self._data = data

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return self._data.shape[1]

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if index.isValid():
            if role == QtCore.Qt.DisplayRole:
                return str(self._data.iloc[index.row(), index.column()])
        return None

    def headerData(self, col, orientation, role):
        if orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole:
            return self._data.columns[col]
        return None


# class to handle threading of datapoints so GUI is responsive
class DataPointsWorkThread(QThread):
	signal = pyqtSignal('PyQt_PyObject')
	signal_pb = pyqtSignal('PyQt_PyObject')
	def __init__(self, data, start_date, end_date, pb_inc, option):
		QThread.__init__(self)
		# create instance of WorkerThread class and pass variables from application class as instance variables
		self.data = data
		self.start_date = start_date
		self.end_date = end_date
		self.pb_inc = pb_inc
		self.option = option

	def run(self):
		# local variables from instance variables for reference convenience
		data = self.data
		start_date = self.start_date
		end_date = self.end_date
		pb_inc = self.pb_inc
		option = self.option

		# initialize datapoints data frame and progress bar
		df = pd.DataFrame()
		pb_update = 0

		# group data into month intervals increasing each item by 1 day
		for date in pd.date_range(start_date, end_date):

			start_date_add_days = date + relativedelta(months=+1,days=-1)
			paid_date = start_date_add_days + relativedelta(months=+2)

			if option == "option2":

				start_date_add_month = date + relativedelta(months=+1)
				start_date_add_months = date + relativedelta(months=+2,days=-1)
				paid_date_add_months = start_date_add_months + relativedelta(months=+2)

			elif option == "option3":

				start_date_add_month = date + relativedelta(months=+1)
				start_date_add_months = date + relativedelta(months=+12,days=-1)
				paid_date_add_months = start_date_add_months + relativedelta(months=+2)

			col1 = date
			col2 = start_date_add_days

			# sum payment data following criteria of allowing for additional 2 months of paid dates
			if option == "option1":

				col3 = round(data[(data.a >= date) & (data.a <= start_date_add_days) & (data.b <= paid_date)].sum()['c'],2)

			elif option == "option2":

				col3 = round(data[(data.a >= date) & (data.a <= start_date_add_days) & (data.b <= paid_date)].sum()['c'],2)

			elif option == "option3":

				col3 = round(data[(data.a >= date) & (data.a <= start_date_add_days) & (data.b <= paid_date)].sum()['c'],2) * 12
			
			if option == "option1":

				df = df.append({'A' : col1 , 'B' : col2, 'C' : col3},ignore_index=True)

			else:

				col4 = round(data[(data.a >= start_date_add_month) & (data.a <= start_date_add_months) & (data.b <= paid_date_add_months)].sum()['c'],2)
				col5 = col3 - col4
				df = df.append({'A' : col1 , 'B' : col2, 'C' : col3, 'D' : col4, 'E' : col5},ignore_index=True)

			# update progress
			pb_update = pb_update + (100/pb_inc)
			self.signal_pb.emit(pb_update)


		# find index of maximum and corresponding date refactor == True is redundant isChecked is Boolean
		if option == "option1":
			index_max_C = df[(df.C == df.max()['C'])].index.tolist()
		else:
			index_max_C = df[(df.E == df.max()['E'])].index.tolist()

		list_date_max_C = df['A'].iloc[index_max_C].tolist()
		list_dollars_max_C = df['C'].iloc[index_max_C].tolist()

		# clear dataframe
		df = pd.DataFrame() 

		# reconstruct dataframe respective to maximum date
		date_max_C = list_date_max_C[0]
		dollars_max_C = list_dollars_max_C[0]

		DOI = (date_max_C + relativedelta(months=+1)).strftime("%#m/%#d/%Y")
		DOI_dollars = '${:,.2f}'.format(dollars_max_C)
		date_max_C_end = date_max_C + relativedelta(months=+3,days=-1)

		for its in range(36):
			df = df.append({'A': str(date_max_C).replace('00:00:00',''), 'B': str(date_max_C + relativedelta(months=+1,days=-1)).replace('00:00:00',''), 'C': round(data[(data.a >= date_max_C) & (data.a <= date_max_C + relativedelta(months=+1,days=-1)) & (data.b <= date_max_C_end)].sum()['c'],2)},ignore_index=True)
			# update progress
			pb_update = pb_update + (100/pb_inc)
			self.signal_pb.emit(pb_update)
			date_max_C = date_max_C + relativedelta(months=-1)

		# sort new dataframe and reset index
		df = df.sort_values(by="A")
		df = df.reset_index(drop=True)

		# define x and y outputs
		x = np.arange(1,37)
		y = np.array(df['C'].tolist())

		# drop and readd A and B with formatting
		A_format, B_format = [], []
		a = np.array(df['A'].tolist())
		b = np.array(df['B'].tolist())
		for each in a:
			A_format.append(pd.to_datetime(each).strftime("%#m/%#d/%Y"))
		df.drop(columns=['A'])
		df['A'] = A_format

		for each in b:
			B_format.append(pd.to_datetime(each).strftime("%#m/%#d/%Y"))
		df.drop(columns=['A'])
		df['B'] = B_format

		# replace with final
		C_format = []
		for each in y:
			C_format.append('${:,.2f}'.format(each))
		df.drop(columns=['C'])
		df['C'] = C_format
		
		# create a dictionary of variables to pass to display
		dp_output = {
			"df":df,
			"x":x,
			"y":y,
			"DOI":DOI,
			"DOI_dollars":DOI_dollars
		}

		# emitting a pyqtSignal named display_output with output dictionary data
		self.signal.emit(dp_output)


# class to handle threading of regression so GUI is responsive
class RegressionWorkThread(QThread):
	signal = pyqtSignal('PyQt_PyObject')

	def __init__(self, x, y, df, max_segments, max_iter, isnt_discretized):
		QThread.__init__(self)
		# create instance of WorkerThread class and pass variables from application class as instance variables
		self.x = x
		self.y = y
		self.df = df
		self.max_segments = max_segments
		self.max_iter = max_iter
		self.isnt_discretized = isnt_discretized

	def run(self):
		# local variables from instance variables for reference convenience
		x = self.x
		y = self.y
		df = self.df
		max_segments = self.max_segments
		max_iter = self.max_iter
		isnt_discretized = self.isnt_discretized

		# reduce df if user has already populated df, selected new option, and ran again
		if len(df.columns) > 3:
			df = df[['Date Range Start','Date Range End','Sum']]

		def my_obj(x):
			l = y.mean()*0.001 # penalty parameter
			f = np.zeros(x.shape[0])
			for i, j in enumerate(x):
				my_pwlf.fit(j[0])
				f[i] = my_pwlf.ssr + (l*j[0])
			return f

		# initialize piecewise linear fit with your x and y data
		my_pwlf = pwlf.PiecewiseLinFit(x, y)

		# define the lower and upper bound for the number of line segements
		bounds = [{'name': 'var_1', 'type': 'discrete', 'domain': np.arange(2, max_segments + 1)}]

		np.random.seed(12121)

		myBopt = BayesianOptimization(my_obj, domain=bounds, model_type='GP',
		                    initial_design_numdata=10,
		                    initial_design_type='latin',
		                    exact_feval=True, verbosity=False,
		                    verbosity_model=False)

		# perform the bayesian optimization to find the optimum number of line segments
		myBopt.run_optimization(max_iter=max_iter, verbosity=False)

		# perform the fit for the optimum
		my_pwlf.fit(myBopt.x_opt)

		# generate regression model and prepare variables and stats for df 
		if isnt_discretized: # time recode is continuous without discretization, all explanatory variables vary
			
			# predict for the determined points
			xHat = np.linspace(min(x), max(x), num=3501) # stretch linespace so segments are not jagged
			yHat = my_pwlf.predict(xHat)

			# calculate n
			n = len(x)

			# get model parameters
			beta = my_pwlf.beta

			# calculate k
			k = len(beta)

			# calculate the standard errors associated with each beta parameter
			se = my_pwlf.standard_errors()

			# calculate t-value
			t = beta / se

			# calculate the p-values
			pvalues = my_pwlf.p_values()

			# calculate r-squared, multiple r, and r-squared adjusted
			# because k includes y-intercept: n-(k+1) => (n-k) for r_sq_adj, mse, and dof  ,  (k) => (k-1) for dof and msr
			r_sq = my_pwlf.r_squared()
			r_mult = math.sqrt(r_sq)
			r_sq_adj = 1 - ((n - 1) / (n - k) * (1 - r_sq)) 

			# calculate sums of squares, means of squares, and standard error
			fit_breaks = my_pwlf.fit_breaks
			ybar = np.ones(my_pwlf.n_data) * np.mean(my_pwlf.y_data)
			ydiff = my_pwlf.y_data - ybar

			sst = np.dot(ydiff, ydiff)
			sse = my_pwlf.fit_with_breaks(fit_breaks) 
			ssr = (sst - sse)

			mse = sse / (n - k)
			msr = ssr / (k - 1)

			S = math.sqrt(mse)

			# calculate F-statistic
			Fstat = (msr / mse)
			
			# calculate degrees of freedom (regression, residual/errors, and total)
			dof = [(k - 1),(n - k),(n - 1)] 

			# populate yHats array unique to pwlf
			yHat_values, yHat_index = [], 0

			for yHats in range(1,37):
				yHat_values.append("${:,.2f}".format(yHat[yHat_index]))
				yHat_index += 100

			# construct independent variables dataframe
			# construct the regression matrix
			A = np.zeros((n, my_pwlf.n_parameters))
			A[:, 0] = 1.0
			A[:, 1] = x - my_pwlf.fit_breaks[0]

			for i in range(my_pwlf.n_segments-1):
			    int_locations = x > my_pwlf.fit_breaks[i+1]
			    if sum(int_locations) > 0:
			        int_index = np.argmax(int_locations)
			        A[int_index:, i+2] = x[int_index:] - my_pwlf.fit_breaks[i+1]

			# transform regression matrix to a dataframe with structure columns = yint, x1, x2, ..., xn
			B = list(map(list,zip(*A)))

			# construct independent variables dataframe
			df_variables = pd.DataFrame()

			for arrays in range(len(A[0])):
				df_variables.insert(loc=arrays,column="col:"+str(arrays),value=B[arrays])
			
			# drop y-intercept column
			df_variables.drop(df_variables.columns[0], axis=1, inplace=True)



		else: # time recode is continuous with discretization, one explanatory variables varies while others held constant


			# discretize breakpoints
			breaks = my_pwlf.fit(myBopt.x_opt)
			breaks_int = []
			for breakpoint in breaks:
				breaks_int.append(round(breakpoint,0))

			# construct regression matrix
			result = [];
			template = [0] * ( len(breaks_int) - 1 ) # create a 0-initialized array of the length of the number of segments
			cursorPosition = 0;
			cursorValue = 1;
			cursorMax = breaks_int[cursorPosition+1]

			for row in range( int(breaks_int[-1]) ):
				thisrow = template.copy()
				# change the value for this row
				thisrow[cursorPosition] = cursorValue
				result.append(thisrow)
				# refer to the result to build on next time
				template = thisrow
				# move the cursor and reset its values
				if (cursorValue >= cursorMax):
					cursorPosition += 1
					if cursorPosition >= (len(breaks_int) - 1):
						break
					cursorValue = 1
					cursorMax = breaks_int[cursorPosition+1] - breaks_int[cursorPosition]
				else:
					cursorValue += 1

			# transpose A so the form is correct
			result = list(map(list, zip(*result)))

			# add intercept row
			result.append([1]*len(result[0]))

			# transpose to regression matrix form
			A = (np.array(result)).T

			# calculate beta and  sse
			# note: y-intercept is last value in beta
			beta, sse, rank, s = np.linalg.lstsq(A, y, rcond=None)

			# predict for the determined points
			xHat = np.linspace(min(x), max(x), num=36, endpoint=True)
			yHat = np.dot(A,beta)

			# calculate n
			n = len(x)

			# calculate k
			k = len(beta)

			# calculate residuals
			e = yHat - y

			# calculate variance
			variance = np.dot(e, e) / (n - k)

			# calculate se
			se = np.sqrt(variance * (np.linalg.inv(np.dot(A.T,A)).diagonal()))

			# calculate t-value
			t = beta / se

			# calculate p-value
			pvalues = 2.0 * stats.t.sf(np.abs(t), df=n-k-1)

			# calculate sums of squares, means of squares, and standard error
			# because k includes y-intercept: n-(k+1) => (n-k) for r_sq_adj, mse, and dof  ,  (k) => (k-1) for dof and msr
			ybar = np.ones(n) * np.mean(y)
			ydiff = y - ybar

			sst = np.dot(ydiff, ydiff)
			sse = sse[0]
			ssr = (sst - sse)

			mse = sse / (n - k)
			msr = ssr / (k - 1)

			S = math.sqrt(mse)

			# calculate F-statistic
			Fstat = (msr / mse)
			
			# calculate degrees of freedom (regression, residual/errors, and total)
			dof = [(k - 1),(n - k),(n - 1)]

			# calculate r-squared, multiple r, and r-squared adjusted
			r_sq = 1.0 - (sse / sst)
			r_mult = math.sqrt(r_sq)
			r_sq_adj = 1 - ((n - 1) / (n - k) * (1 - r_sq)) 

			# construct independent variables dataframe
			df_variables = pd.DataFrame()

			loc = 0
			colnum = 0
			for arrays in result:
				col = "col" + str(colnum)
				df_variables.insert(loc=loc,column=col,value=arrays)
				loc += 1
				colnum += 1

			# drop y-intercept column
			df_variables.drop(df_variables.columns[-1], axis=1, inplace=True)

			# populate yHats array unique to this option
			yHat_values = []
			for yHats in yHat:
				yHat_values.append("${:,.2f}".format(yHats))

			# in discrete calcs y-intercept parameters are listed last as opposed to first in pwlf, so we need to reorder
			new_beta, new_se, new_t, new_pvalues = [], [], [], []
			new_beta.append(beta[-1])
			new_se.append(se[-1])
			new_t.append(t[-1])
			new_pvalues.append(pvalues[-1])
			for i in range(0,k-1):
				new_beta.append(beta[i])
				new_se.append(se[i])
				new_t.append(t[i])
				new_pvalues.append(pvalues[i])
			beta = new_beta
			se = new_se
			t = new_t
			pvalues = new_pvalues

		# complete dataframes
		# insert yHats into df
		df.insert(loc=3,column='col4',value=yHat_values)

		# insert ind variables into df
		loc = 4

		for columns in df_variables:
			df.insert(loc=loc,column='x-'+str(loc-3),value=df_variables[columns])
			loc += 1

		# build summary statistics dataframe
		regres_stats_labels = ["Multiple R","R Square","Adjusted R Square","Standard Error","Observations"]
		regress_stats = ["{:0.2f}".format(r_mult),"{:0.2f}".format(r_sq),"{:0.2f}".format(r_sq_adj),"{:0.2f}".format(S),n]

		df_regress_stats = pd.DataFrame({"Regression":regres_stats_labels,"Statistics":regress_stats})

		# build ANOVA dataframe
		SS = ["{:0.2f}".format(ssr),"{:0.2f}".format(sse),"{:0.2f}".format(sst)]
		MS = ["{:0.2f}".format(msr),"{:0.2f}".format(mse),'']
		F = ["{:0.2f}".format(Fstat),'','']
		anova_labels = ['Regression','Residual','Total']

		df_anova = pd.DataFrame({'':anova_labels,'df':dof,'SS':SS,'MS':MS,'F':F})

		# build coefficients dataframe
		df_coef_labels = []
		df_coef_labels.append('Y-Intercept')

		for i in range(1,k):
			df_coef_labels.append('x-'+str(i))

		beta_format, se_format, t_format, pvalues_format = [], [], [], []

		for i in range(k):
			roundbeta, roundse, roundt, roundp = "{:0.2f}".format(beta[i]), "{:0.2f}".format(se[i]), "{:0.2f}".format(t[i]), "{:0.2f}".format(pvalues[i])
			beta_format.append(roundbeta)
			se_format.append(roundse)
			t_format.append(roundt)
			pvalues_format.append(roundp)

		df_coef = pd.DataFrame({'':df_coef_labels,"Coefficients":beta_format,"Standard Error":se_format,"t Stat":t_format,"P-value":pvalues_format})

		# plot the results and save as a temporary file to be overwritten each iterations
		plt.figure()
		plt.plot(x, y, '-')
		plt.plot(xHat, yHat, 'r--')

		# # provide number of segments from model
		num_segments = str(myBopt.x_opt).replace("[","").replace(".]","")

		# provide function value from model
		func_value = "{:0.2f}".format(myBopt.fx_opt)

		# create a dictionary of variables to pass to display
		regression_output = {
			"num_segments":num_segments,
			"func_value":func_value,
			"df_regress_stats":df_regress_stats,
			"df_anova":df_anova,
			"df_coef":df_coef,
			"x":x,
			"y":y,
			"xHat":xHat,
			"yHat":yHat,
			"df":df,
			"plt":plt,
		}

		# emitting a pyqtSignal with output dictionary data
		self.signal.emit(regression_output)


# main window class
class DPR(QtWidgets.QMainWindow): 
	def __init__(self, parent=None):
		# call the parent class's constructor
		QtWidgets.QMainWindow.__init__(self, parent)
		self.ui = Ui_MainWindow()
		self.ui.setupUi(self)
		self.ui.pushButton_1.clicked.connect(self.select_file)
		self.ui.pushButton_2.clicked.connect(self.run_datapoints)
		self.ui.dateEdit_1.setDateTime(QtCore.QDateTime.currentDateTime())
		self.ui.dateEdit_2.setDateTime(QtCore.QDateTime.currentDateTime())
		self.ui.dateEdit_1.dateChanged.connect(self.update_date)
		self.ui.radioButton_1.setChecked(True)
		self.ui.radioButton_4.setChecked(True)
		self.ui.pushButton_3.clicked.connect(self.run_regression)
		self.ui.lineEdit_3.setText("5")
		self.ui.lineEdit_4.setText("10")
		self.ui.pushButton_4.clicked.connect(self.write_excel)
		self.ui.graphicsView_1.hide()
		self.ui.graphicsView_2.hide()
		self.MessageBox = ctypes.windll.user32.MessageBoxW


	# after first date edit is changed, update second date edit to be a year later
	def update_date(self):
		get_date = self.ui.dateEdit_1.date().toString("yyyy-M-d")
		new_datetime = pd.to_datetime(get_date) + relativedelta(months=+12)
		change_datetime = QtCore.QDateTime.fromString(str(new_datetime), "yyyy-M-d hh:mm:ss")
		self.ui.dateEdit_2.setDateTime(change_datetime)


	# check if file is selected
	def select_file(self):
		filename, _ = QtWidgets.QFileDialog.getOpenFileName(None, "Select File", "","Text Files (*.txt)")
		if filename:
			# outputs
			self.ui.lineEdit_1.setText(filename)
			self.filename = filename


	def run_datapoints(self):		
		delimiter = str(self.ui.comboBox_1.currentText())
		has_headers = self.ui.checkBox_1.isChecked()

		if self.ui.lineEdit_1.text() == "":
			self.MessageBox(None, "No file selected.", "File Error", 0)
			return

		try:
			data = self.prepare_data(delimiter, has_headers)
		except pd.errors.EmptyDataError:
			self.MessageBox(None, "No data in file.", "Empty Data Error", 0)
			return

		if data is 0:
			self.MessageBox(None, "Problem reading file. Check header declaration.", "Attribute Error", 0)
		elif data is 1:
			self.MessageBox(None, "Column 1 should be date type.", "Attribute Error", 0)
			return
		elif data is 2:
			self.MessageBox(None, "Column 2 should be date type.", "Attribute Error", 0)
			return
		elif data is 3:
			self.MessageBox(None, "Column 3 should be currency.", "Attribute Error", 0)
			return

		# disable calculate button
		self.ui.pushButton_2.setEnabled(False)

		start_date = pd.to_datetime(self.ui.dateEdit_1.date().toString("M/d/yyyy"))
		end_date = pd.to_datetime(self.ui.dateEdit_2.date().toString("M/d/yyyy"))
		pb_inc = (end_date - start_date).days + 36 #number of items in the 2 loops in datapoints fxs

		if self.ui.radioButton_1.isChecked():
			option = "option1"
		elif self.ui.radioButton_2.isChecked():
			option = "option2"
		else: option = "option3"

		self.worker_thread = DataPointsWorkThread(data, start_date, end_date, pb_inc, option)
		self.worker_thread.signal.connect(self.display_datapoints)
		self.worker_thread.signal_pb.connect(self.update_progressbar)
		self.worker_thread.start()


	def update_progressbar(self, pb_update):
		self.ui.progressBar_1.setValue(pb_update)


	# construct raw data dataframe from file data
	def prepare_data(self, delimiter, has_headers):

		if delimiter == "Tab Delimited":
			sep = "\t"
		elif delimiter == 'Comma Delimited':
			sep = ","
		elif delimiter == 'Pipe Delimited':
			sep = "|"

		if has_headers: # data file has headers
			try:
				data = pd.read_csv(self.filename, skiprows=0, sep=sep, header=None)
			except AttributeError:
				return 0
		else: # data file does not have headers
			try:
				data = pd.read_csv(self.filename, sep=sep, header=None)
			except AttributeError:
				return 0

		data.columns = ["a", "b", "c"]

		try:
			data['a'] = pd.to_datetime(data['a'])
		except ValueError:
			return 1

		try:
			data['b'] = pd.to_datetime(data['b'])
		except ValueError:
			return 2

		try:
			data['c'] = data['c'].replace('[\$,]', '', regex=True).astype(float)
		except ValueError:
			return 3

		return data


	# display dataframe with more specific headers
	def display_datapoints(self, dp_output):
		self.df = dp_output["df"]
		self.x = dp_output["x"]
		self.y = dp_output["y"]
		self.DOI = dp_output["DOI"]
		self.DOI_dollars = dp_output["DOI_dollars"]

		df = self.df
		x = self.x
		y = self.y
		DOI = self.DOI
		DOI_dollars = self.DOI_dollars

		# update progress bar
		self.ui.progressBar_1.setValue(100) 

		# graph plot
		self.ui.graphicsView_1.canvas.ax.clear() # clear if already drawn
		self.ui.graphicsView_1.canvas.ax.plot(x, y, '-', color="orange")
		self.ui.graphicsView_1.canvas.ax.set_xlabel('Time', fontsize = 8.0) 
		self.ui.graphicsView_1.canvas.ax.set_ylabel('Sum', fontsize = 8.0) 
		self.ui.graphicsView_1.canvas.draw() # needed for the first drawing
		for ytick in self.ui.graphicsView_1.canvas.ax.get_yticklabels(): 
			ytick.set_fontsize(8.0)
		for xtick in self.ui.graphicsView_1.canvas.ax.get_xticklabels(): 
			xtick.set_fontsize(8.0)
		labels = [item.get_text() for item in self.ui.graphicsView_1.canvas.ax.get_yticklabels()]
		try:
			self.ui.graphicsView_1.canvas.ax.set_yticklabels(['${:,}'.format(int(label)) for label in labels])
		except ValueError:
			self.MessageBox(None, "Error plotting data with date ranges provided.", "Value Error", 0)
			return
		self.ui.graphicsView_1.canvas.draw() # needed for updated figures
		self.ui.graphicsView_1.show()
		self.ui.lineEdit_2.setText(DOI)
		self.ui.lineEdit.setText(DOI_dollars)

		# enable calculate button
		self.ui.pushButton_2.setEnabled(True)

		# table
		df.columns = ["Date Range Start", "Date Range End", "Sum"]
		model = PandasModel(df)
		self.ui.tableView_1.setModel(model)


	def run_regression(self):
		x = self.x
		y = self.y
		try:
			df = self.df
		except AttributeError:
			self.MessageBox(None, "No datapoints have been calculated.", "Attribute Error", 0)
			return

		max_segments = int(self.ui.lineEdit_3.text())
		max_iter = int(self.ui.lineEdit_4.text())
		isnt_discretized = self.ui.radioButton_4.isChecked()

		# disable calculate button
		self.ui.pushButton_3.setEnabled(False)

		self.ui.progressBar_2.setRange(0,0)
		self.worker_thread = RegressionWorkThread(x, y, df, max_segments, max_iter, isnt_discretized)
		self.worker_thread.signal.connect(self.display_regression)
		self.worker_thread.start()


	def display_regression(self, regression_output):
		# instance variables for excel output
		self.df_output = regression_output["df"]
		self.df_regress_stats = regression_output["df_regress_stats"]
		self.df_anova = regression_output["df_anova"]
		self.df_coef = regression_output["df_coef"]
		self.plt = regression_output["plt"]

		# update progress bar
		self.ui.progressBar_2.setRange(0,100)
		self.ui.progressBar_2.setValue(100)

		# graph plot
		self.ui.graphicsView_2.canvas.ax.clear() # clear if already drawn
		self.ui.graphicsView_2.canvas.ax.plot(regression_output["x"], regression_output["y"],'-', color="orange")
		self.ui.graphicsView_2.canvas.ax.plot(regression_output["xHat"], regression_output["yHat"],'--', color="royalblue")
		self.ui.graphicsView_2.canvas.ax.set_xlabel('Time', fontsize = 8.0) 
		self.ui.graphicsView_2.canvas.ax.set_ylabel('Sum', fontsize = 8.0) 
		self.ui.graphicsView_2.canvas.draw()
		for ytick in self.ui.graphicsView_2.canvas.ax.get_yticklabels(): 
			ytick.set_fontsize(8.0)
		for xtick in self.ui.graphicsView_2.canvas.ax.get_xticklabels(): 
			xtick.set_fontsize(8.0)
		labels = [item.get_text() for item in self.ui.graphicsView_2.canvas.ax.get_yticklabels()]
		try:
			self.ui.graphicsView_2.canvas.ax.set_yticklabels(['${:,}'.format(int(label)) for label in labels])
		except ValueError:
			self.MessageBox(None, "Error plotting data with date ranges provided.", "Value Error", 0)
			return
		self.ui.graphicsView_2.canvas.draw()
		self.ui.graphicsView_2.show()

		# provide number of segments from model
		self.ui.lineEdit_5.setText(regression_output["num_segments"])

		# provide function value from model
		self.ui.lineEdit_6.setText(regression_output["func_value"])

		# display regression statistics
		model_regress_stats = PandasModel(regression_output["df_regress_stats"])
		self.ui.tableView_2.setModel(model_regress_stats)

		# display ANOVA statistics
		model_anova = PandasModel(regression_output["df_anova"])
		self.ui.tableView_3.setModel(model_anova)

		# display coefficients table
		model_coef = PandasModel(regression_output["df_coef"])
		self.ui.tableView_4.setModel(model_coef)

		# enable calculate button
		self.ui.pushButton_3.setEnabled(True)


	def write_excel(self):
		try:
			df_output = self.df_output
		except AttributeError:
			self.MessageBox(None, "No datapoints have been calculated.", "Attribute Error", 0)
			return

		df_regress_stats = self.df_regress_stats
		df_anova = self.df_anova
		df_coef = self.df_coef
		plt = self.plt

		# prompt for save file path and name, default is desktop
		desktop = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')
		fileout, _ = QtWidgets.QFileDialog.getSaveFileName(None,"Save File", desktop, "Excel File (*xlsx)")

		if fileout:
			# append .xlsx to file name 
			if fileout.endswith(".xlsx"):
				fileout = fileout
			else: fileout += ".xlsx"

			# split path into basepath and file name
			head, tail = os.path.split(fileout)

			# save figure to be overwritten each iteration in project folder
			plt.savefig("myplot.png", dpi = 100)

			# rename column headers
			df_output = df_output.rename(index=str, columns={"C": "Sum", "col4":"Estimate (y-prediction)"})

			# write dataframes to an input excel file to be overwritten each iteration in project folder
			num_columns = len(df_output.columns)
			writer = pd.ExcelWriter('input.xlsx',engine='xlsxwriter')
			df_output.to_excel(writer,sheet_name='Sheet1',index=False)
			df_regress_stats.to_excel(writer,sheet_name='Sheet1',startrow=30,startcol=num_columns+1, index=False)
			df_anova.to_excel(writer,sheet_name="Sheet1",startrow=38,startcol=num_columns+1,index=False)
			df_coef.to_excel(writer,sheet_name='Sheet1',startrow=43,startcol=num_columns+1, index=False)
			writer.save()

			# open input excel file, do formatting, add plot, save file as output
			wb = openpyxl.load_workbook("input.xlsx")
			ws = wb.active
			ws.merge_cells('A1:B1')
			ws.cell(row=1,column=1).value = "Service Date Range"
			img = openpyxl.drawing.image.Image("myplot.png")
			ws.add_image(img, 'M1')
			wb.save(fileout)
			wb.close()

# Create QT Application
app = QtWidgets.QApplication(sys.argv)

# Create and show UI View
myInstance = DPR()
myInstance.show()

# Start the application, log any errors as exit code
sys.exit(app.exec_())